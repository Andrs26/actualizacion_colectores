import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.http import JsonResponse
from django.db.models.functions import Cast
from django.db.models import CharField
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from actualizacion.models import Cliente, ClienteDatos, ClienteContacto
import json
import pandas as pd
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.contrib.auth.models import User
from django.contrib.auth.models import Group
import json
from datetime import datetime, timedelta
import math
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

@login_required
def dashboard_home(request):
    user_groups = request.user.groups.values_list('name', flat=True)

    total_clientes = Cliente.objects.count()
    total_actualizados = Cliente.objects.filter(estado='Actualizado').count()
    total_no_ubicados = Cliente.objects.filter(estado='No Ubicado').count()
    total_pendientes = Cliente.objects.filter(estado='Pendiente').count()

    # Calcular avance (progreso) y determinar color de barra
    if total_clientes:
        avance = round((total_actualizados + total_no_ubicados) * 100 / total_clientes)
    else:
        avance = 0
    if avance >= 70:
        barra_color = 'bg-success'
    elif avance >= 30:
        barra_color = 'bg-primary'
    elif avance >= 1:
        barra_color = 'bg-danger'
    else:
        barra_color = 'bg-secondary'

    context = {
        'is_admin': request.user.is_superuser or 'admin' in user_groups,
        'is_colector': 'colector' in user_groups,
        'total_clientes': total_clientes,
        'total_actualizados': total_actualizados,
        'total_no_ubicados': total_no_ubicados,
        'total_pendientes': total_pendientes,
        'avance': avance,
        'barra_color': barra_color,
    }

    chart_data = {
        'labels': json.dumps(['Actualizados', 'No Ubicados', 'Pendientes']),
        'values': json.dumps([total_actualizados, total_no_ubicados, total_pendientes]),
    }
    context.update(chart_data)

    # Filtros para chart agrupado por fecha, usuario y estado
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    estado = request.GET.get('estado')
    usuario_id = request.GET.get('usuario_id')

    filtros = {}
    # Usar __range directamente porque updated_at es DateField
    if start_date and end_date:
        filtros['updated_at__range'] = [start_date, end_date]
    if estado:
        filtros['estado'] = estado
    if usuario_id:
        filtros['usuario_asignado__id'] = usuario_id

    # Agrupado por fecha (manual truncation for SQLite), usuario y estado
    datos = Cliente.objects.annotate(
        fecha=Cast('updated_at', output_field=CharField())
    ).filter(
        estado__in=['Actualizado', 'No Ubicado'],
        **filtros
    ).values(
        'fecha', 'usuario_asignado__username', 'estado'
    ).annotate(
        total=Count('id')
    ).order_by('fecha')

    # Convertir fecha a string solo con la parte de fecha (YYYY-MM-DD)
    datos_list = []
    for d in datos:
        d = dict(d)
        if d['fecha']:
            d['fecha'] = d['fecha'][:10]
        else:
            d['fecha'] = ''
        datos_list.append(d)

    usuarios = User.objects.exclude(username='admin')

    context['datos_chart'] = json.dumps(datos_list, default=str)
    context['usuarios'] = usuarios

    # Lista de usuarios para el filtro clásico
    usuarios_list = usuarios.values('id', 'username', 'first_name', 'last_name')
    context['usuarios_list'] = usuarios_list

    # ---- COLORES Y USUARIOS-ESTADOS PARA EL CHART ----
    import colorsys
    def generar_colores_hex(n):
        colores = []
        for i in range(n):
            h = i / n
            r, g, b = colorsys.hsv_to_rgb(h, 0.6, 0.85)
            colores.append(f'rgba({int(r*255)}, {int(g*255)}, {int(b*255)}, 0.7)')
        return colores

    # Generar la lista única de etiquetas "usuario - estado"
    usuarios_estados = list(set(f"{d['usuario_asignado__username']} - {d['estado']}" for d in datos_list))
    context['usuarios_estados'] = json.dumps(usuarios_estados)
    context['colores_chart'] = json.dumps(generar_colores_hex(len(usuarios_estados)))

    # --- RESUMEN USUARIOS PARA TABLA PROJECTS ---
    usuarios_proyectos = Cliente.objects.values_list('usuario_asignado', flat=True).distinct()
    resumen_usuarios = []
    for usuario_id in usuarios_proyectos:
        if not usuario_id:
            continue
        user_obj = User.objects.filter(id=usuario_id).first()
        total = Cliente.objects.filter(usuario_asignado=usuario_id).count()
        actualizados = Cliente.objects.filter(usuario_asignado=usuario_id, estado='Actualizado').count()
        no_ubicados = Cliente.objects.filter(usuario_asignado=usuario_id, estado__iexact='No Ubicado').count()
        porcentaje = int((actualizados / total) * 100) if total > 0 else 0
        if porcentaje >= 70:
            color = 'bg-gradient-success'
        elif porcentaje >= 30:
            color = 'bg-gradient-info'
        elif porcentaje > 0:
            color = 'bg-gradient-danger'
        else:
            color = 'bg-gradient-secondary'
        resumen_usuarios.append({
            'first_name': user_obj.first_name if user_obj else '',
            'last_name': user_obj.last_name if user_obj else '',
            'actualizados': actualizados,
            'no_ubicados': no_ubicados,
            'total': total,
            'porcentaje': porcentaje,
            'color': color,
        })
    context['resumen_usuarios'] = resumen_usuarios

    # --- NUEVO: progreso_autores para tabla "Authors table" ---
    from django.contrib.auth.models import Group
    from datetime import timedelta, date

    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())  # Lunes de esta semana

    # Filtrar solo usuarios del grupo "colector"
    try:
        grupo_colectores = Group.objects.get(name="colector")
        usuarios = grupo_colectores.user_set.all()
    except Group.DoesNotExist:
        usuarios = User.objects.none()

    progreso_por_usuario = []

    for user in usuarios:
        total_asignados = Cliente.objects.filter(usuario_asignado=user).count()
        total_procesados = Cliente.objects.filter(
            usuario_asignado=user,
            updated_at__range=(inicio_semana, hoy),
            estado__in=['Actualizado', 'No Ubicado']
        ).count()
        dias_transcurridos = max((hoy - inicio_semana).days + 1, 1)
        promedio_diario = round(total_procesados / dias_transcurridos) if dias_transcurridos > 0 else 0
        pendientes = total_asignados - total_procesados
        if promedio_diario > 0:
            tiempo_estimado_dias = round(pendientes / promedio_diario)
            tiempo_estimado_dias = max(tiempo_estimado_dias, 0)
        else:
            tiempo_estimado_dias = None
        progreso_por_usuario.append({
            'usuario': f"{user.first_name} {user.last_name}",
            'promedio_diario': promedio_diario,
            'pendientes': pendientes,
            'tiempo_estimado': tiempo_estimado_dias if tiempo_estimado_dias is not None else "N/A",
        })

    context['progreso_autores'] = progreso_por_usuario

    return render(request, 'dashboard/index.html', context)

# Nueva vista AJAX para Chart.js: Avance por fecha, usuario y estado
from django.http import JsonResponse
from django.db.models import Count
from datetime import datetime, date
from django.contrib.auth.decorators import login_required

@login_required
def dashboard_chart_data(request):
    filtros = {}
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado')
    usuario = request.GET.get('usuario')

    if fecha_inicio and fecha_fin:
        filtros['updated_at__range'] = [fecha_inicio, fecha_fin]
    elif fecha_inicio:
        filtros['updated_at'] = fecha_inicio

    if estado:
        filtros['estado'] = estado
    if usuario:
        filtros['usuario_asignado'] = usuario

    queryset = Cliente.objects.filter(**filtros).values('updated_at__date', 'estado', 'usuario_asignado').annotate(cantidad=Count('id')).order_by('updated_at__date')

    datos_por_fecha = {}
    for fila in queryset:
        fecha = fila['updated_at__date'].strftime("%d-%b")
        estado_val = fila['estado']
        usuario_val = str(fila['usuario_asignado']) if fila['usuario_asignado'] else 'No asignado'
        if fecha not in datos_por_fecha:
            datos_por_fecha[fecha] = {}
        if estado_val not in datos_por_fecha[fecha]:
            datos_por_fecha[fecha][estado_val] = {}
        if usuario_val not in datos_por_fecha[fecha][estado_val]:
            datos_por_fecha[fecha][estado_val][usuario_val] = 0
        datos_por_fecha[fecha][estado_val][usuario_val] += fila['cantidad']

    fechas = sorted(datos_por_fecha.keys())
    estados = set()
    usuarios = set()

    for est_dict in datos_por_fecha.values():
        for est, user_dict in est_dict.items():
            estados.add(est)
            usuarios.update(user_dict.keys())

    estados = sorted(estados)
    usuarios = sorted(usuarios)

    datasets = []
    for estado_val in estados:
        for usuario_val in usuarios:
            data = []
            for fecha in fechas:
                cant = datos_por_fecha.get(fecha, {}).get(estado_val, {}).get(usuario_val, 0)
                data.append(cant)
            datasets.append({
                "label": f"{usuario_val} - {estado_val}",
                "data": data
            })

    return JsonResponse({
        "labels": fechas,
        "datasets": datasets
    })

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


# AJAX endpoints para búsqueda y paginación de clientes por tab
from django.template.loader import render_to_string

from django.views.decorators.http import require_GET

@login_required
@require_GET
def clientes_pendientes(request):
    from django.core.paginator import Paginator
    q = request.GET.get('q', '')
    page = request.GET.get('page', 1)
    clientes = Cliente.objects.filter(
        estado='Pendiente',
        nombre_empresa__icontains=q
    )
    paginator = Paginator(clientes, 25)
    page_obj = paginator.get_page(page)
    context = {'clientes': page_obj, 'page_obj': page_obj}
    return JsonResponse({
        'tabla': render_to_string('dashboard/partials_admin/tabla_pendientes.html', context, request=request),
        'paginacion': render_to_string('dashboard/partials_admin/paginacion.html', context, request=request)
    })

@login_required
@require_GET
def clientes_actualizados(request):
    from django.core.paginator import Paginator
    q = request.GET.get('q', '')
    page = request.GET.get('page', 1)
    clientes = Cliente.objects.filter(
        estado='Actualizado',
        nombre_empresa__icontains=q
    )
    paginator = Paginator(clientes, 25)
    page_obj = paginator.get_page(page)
    context = {'clientes': page_obj, 'page_obj': page_obj}
    return JsonResponse({
        'tabla': render_to_string('dashboard/partials_admin/tabla_actualizados.html', context, request=request),
        'paginacion': render_to_string('dashboard/partials_admin/paginacion.html', context, request=request)
    })

@login_required
@require_GET
def clientes_no_ubicados(request):
    from django.core.paginator import Paginator
    q = request.GET.get('q', '')
    page = request.GET.get('page', 1)
    clientes = Cliente.objects.filter(
        estado='No Ubicado',
        nombre_empresa__icontains=q
    )
    paginator = Paginator(clientes, 25)
    page_obj = paginator.get_page(page)
    context = {'clientes': page_obj, 'page_obj': page_obj}
    return JsonResponse({
        'tabla': render_to_string('dashboard/partials_admin/tabla_no_ubicados.html', context, request=request),
        'paginacion': render_to_string('dashboard/partials_admin/paginacion.html', context, request=request)
    })

@login_required
def detalle_cliente_admin(request, cliente_id):
    user_groups = request.user.groups.values_list('name', flat=True)
    cliente = Cliente.objects.get(id=cliente_id)
    datos = getattr(cliente, 'datos', None)
    contactos = cliente.contactos.all()
    
    return render(request, 'dashboard/detalle_cliente.html', {
        'cliente': cliente,
        'datos': datos,
        'contactos': contactos,
        'is_admin': request.user.is_superuser or 'admin' in user_groups,
        'is_colector': 'colector' in user_groups,
    })

@login_required
def ver_detalle_cliente_admin(request, cliente_id):
    user_groups = request.user.groups.values_list('name', flat=True)
    cliente = Cliente.objects.get(id=cliente_id)
    datos = getattr(cliente, 'datos', None)
    contactos = cliente.contactos.all()
    
    return render(request, 'dashboard/ver_detalle_cliente.html', {
        'cliente': cliente,
        'datos': datos,
        'contactos': contactos,
        'is_admin': request.user.is_superuser or 'admin' in user_groups,
        'is_colector': 'colector' in user_groups,
    })

@login_required
def exportar_home(request):
    user_groups = request.user.groups.values_list('name', flat=True)

    from django.contrib.auth.models import Group
    try:
        grupo_colectores = Group.objects.get(name="colector")
        usuarios = grupo_colectores.user_set.all()
    except Group.DoesNotExist:
        usuarios = []

    return render(request, 'dashboard/exportar.html', {
        'is_admin': request.user.is_superuser or 'admin' in user_groups,
        'is_colector': 'colector' in user_groups,
        'usuarios': usuarios
    })


# --- Exportar clientes filtrados a Excel ---
from django.utils.dateparse import parse_date

@login_required
def exportar_clientes_excel(request):
    fecha_inicio = parse_date(request.GET.get('fecha_inicio')) if request.GET.get('fecha_inicio') else None
    fecha_fin = parse_date(request.GET.get('fecha_fin')) if request.GET.get('fecha_fin') else None
    estado = request.GET.get('estado')
    usuario_id = request.GET.get('usuario_id')

    filtros = {}
    if fecha_inicio and fecha_fin:
        filtros['updated_at__range'] = (fecha_inicio, fecha_fin)
    if estado:
        filtros['estado'] = estado
    if usuario_id:
        filtros['usuario_asignado__id'] = usuario_id

    clientes = Cliente.objects.filter(**filtros).select_related('datos', 'usuario_asignado').prefetch_related('contactos')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = [
        'ID Zoho', 'Empresa', 'Estado', 'Usuario Asignado',
        'Dirección', 'Teléfono', 'Correo', 'Página Web',
        'Sucursales', 'Empleados', 'Productos',
        'Contacto 1 Nombre', 'Contacto 1 Teléfono', 'Contacto 1 Correo',
        'Contacto 2 Nombre', 'Contacto 2 Teléfono', 'Contacto 2 Correo',
        'Contacto 3 Nombre', 'Contacto 3 Teléfono', 'Contacto 3 Correo',
        'Contacto 4 Nombre', 'Contacto 4 Teléfono', 'Contacto 4 Correo',
    ]
    ws.append(headers)

    for cliente in clientes:
        datos = cliente.datos if hasattr(cliente, 'datos') else None
        contactos_list = list(cliente.contactos.all())
        contacto_1 = contactos_list[0] if len(contactos_list) > 0 else None
        contacto_2 = contactos_list[1] if len(contactos_list) > 1 else None
        contacto_3 = contactos_list[2] if len(contactos_list) > 2 else None
        contacto_4 = contactos_list[3] if len(contactos_list) > 3 else None

        row = [
            cliente.id_zoho,
            cliente.nombre_empresa,
            cliente.estado,
            f"{cliente.usuario_asignado.first_name} {cliente.usuario_asignado.last_name}" if cliente.usuario_asignado else '',
            datos.direccion if datos else '',
            datos.telefono if datos else '',
            datos.correo if datos else '',
            datos.pagina_web if datos else '',
            datos.numero_sucursales if datos else '',
            datos.numero_empleados if datos else '',
            datos.principales_productos if datos else '',
            f"{contacto_1.nombre} {contacto_1.apellido}" if contacto_1 else '',
            contacto_1.telefono if contacto_1 else '',
            contacto_1.correo if contacto_1 else '',
            f"{contacto_2.nombre} {contacto_2.apellido}" if contacto_2 else '',
            contacto_2.telefono if contacto_2 else '',
            contacto_2.correo if contacto_2 else '',
            f"{contacto_3.nombre} {contacto_3.apellido}" if contacto_3 else '',
            contacto_3.telefono if contacto_3 else '',
            contacto_3.correo if contacto_3 else '',
            f"{contacto_4.nombre} {contacto_4.apellido}" if contacto_4 else '',
            contacto_4.telefono if contacto_4 else '',
            contacto_4.correo if contacto_4 else '',
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes_exportados.xlsx'
    wb.save(response)
    return response


# --- Exportar clientes + datos a Excel ---
from django.contrib.auth.decorators import login_required

@login_required
def exportar_clientes_datos_excel(request):
    clientes = Cliente.objects.select_related('datos', 'usuario_asignado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes + Datos"

    headers = [
        'ID Zoho', 'Empresa', 'Estado', 'Usuario Asignado',
        'Dirección', 'Teléfono', 'Correo', 'Página Web',
        'Sucursales', 'Empleados', 'Productos',
    ]
    ws.append(headers)

    for cliente in clientes:
        datos = cliente.datos if hasattr(cliente, 'datos') else None
        row = [
            cliente.id_zoho,
            cliente.nombre_empresa,
            cliente.estado,
            f"{cliente.usuario_asignado.first_name} {cliente.usuario_asignado.last_name}" if cliente.usuario_asignado else '',
            datos.direccion if datos else '',
            datos.telefono if datos else '',
            datos.correo if datos else '',
            datos.pagina_web if datos else '',
            datos.numero_sucursales if datos else '',
            datos.numero_empleados if datos else '',
            datos.principales_productos if datos else '',
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes_datos.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_clientes_contactos_excel(request):
    modo = request.GET.get('modo', 'columnas')  # 'filas' o 'columnas'

    clientes = Cliente.objects.prefetch_related('contactos', 'usuario_asignado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes + Contactos"

    if modo == 'filas':
        headers = ['ID Zoho', 'Empresa', 'Estado', 'Usuario Asignado', 'Nombre Contacto', 'Teléfono', 'Correo']
        ws.append(headers)
        for cliente in clientes:
            for contacto in cliente.contactos.all():
                ws.append([
                    cliente.id_zoho,
                    cliente.nombre_empresa,
                    cliente.estado,
                    f"{cliente.usuario_asignado.first_name} {cliente.usuario_asignado.last_name}" if cliente.usuario_asignado else '',
                    f"{contacto.nombre} {contacto.apellido}",
                    contacto.telefono,
                    contacto.correo
                ])
    else:  # columnas
        headers = ['ID Zoho', 'Empresa', 'Estado', 'Usuario Asignado',
                   'Contacto 1 Nombre', 'Contacto 1 Teléfono', 'Contacto 1 Correo',
                   'Contacto 2 Nombre', 'Contacto 2 Teléfono', 'Contacto 2 Correo',
                   'Contacto 3 Nombre', 'Contacto 3 Teléfono', 'Contacto 3 Correo',
                   'Contacto 4 Nombre', 'Contacto 4 Teléfono', 'Contacto 4 Correo']
        ws.append(headers)
        for cliente in clientes:
            contactos = list(cliente.contactos.all())
            c1 = contactos[0] if len(contactos) > 0 else None
            c2 = contactos[1] if len(contactos) > 1 else None
            c3 = contactos[2] if len(contactos) > 2 else None
            c4 = contactos[3] if len(contactos) > 3 else None
            row = [
                cliente.id_zoho,
                cliente.nombre_empresa,
                cliente.estado,
                f"{cliente.usuario_asignado.first_name} {cliente.usuario_asignado.last_name}" if cliente.usuario_asignado else '',
                f"{c1.nombre} {c1.apellido}" if c1 else '', c1.telefono if c1 else '', c1.correo if c1 else '',
                f"{c2.nombre} {c2.apellido}" if c2 else '', c2.telefono if c2 else '', c2.correo if c2 else '',
                f"{c3.nombre} {c3.apellido}" if c3 else '', c3.telefono if c3 else '', c3.correo if c3 else '',
                f"{c4.nombre} {c4.apellido}" if c4 else '', c4.telefono if c4 else '', c4.correo if c4 else '',
            ]
            ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'clientes_contactos_filas.xlsx' if modo == 'filas' else 'clientes_contactos_columnas.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
def importar_home(request):
    user_groups = request.user.groups.values_list('name', flat=True)

    from django.contrib.auth.models import Group
    try:
        grupo_colectores = Group.objects.get(name="colector")
        usuarios = grupo_colectores.user_set.all()
    except Group.DoesNotExist:
        usuarios = []

    return render(request, 'dashboard/importar.html', {
        'is_admin': request.user.is_superuser or 'admin' in user_groups,
        'is_colector': 'colector' in user_groups,
        'usuarios': usuarios
    })

@csrf_exempt
@login_required
def importar_clientes_excel(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if archivo:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                id_zoho = fila[0]
                if not id_zoho:
                    continue
                cliente, creado = Cliente.objects.update_or_create(
                    id_zoho=id_zoho,
                    defaults={
                        'nombre_empresa': fila[1],
                        'estado': fila[2],
                        'observaciones': '',
                    }
                )
                if request.user.is_superuser:
                    cliente.usuario_asignado = request.user
                    cliente.save()

                ClienteDatos.objects.update_or_create(
                    cliente=cliente,
                    defaults={
                        'direccion': fila[4],
                        'telefono': fila[5],
                        'correo': fila[6],
                        'pagina_web': fila[7],
                        'numero_sucursales': fila[8],
                        'numero_empleados': fila[9],
                        'principales_productos': fila[10],
                    }
                )

                for i in range(11, 23, 3):
                    if fila[i]:
                        ClienteContacto.objects.update_or_create(
                            cliente=cliente,
                            correo=fila[i+2],
                            defaults={
                                'nombre': fila[i].split()[0] if fila[i] else '',
                                'apellido': ' '.join(fila[i].split()[1:]) if fila[i] else '',
                                'telefono': fila[i+1],
                                'cargo': ''
                            }
                        )
        return redirect('importar_home')

@csrf_exempt
@login_required
def importar_clientes_datos_excel(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if archivo:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            for fila in ws.iter_rows(min_row=2, values_only=True):
                id_zoho = fila[0]
                if not id_zoho:
                    continue
                cliente, _ = Cliente.objects.update_or_create(
                    id_zoho=id_zoho,
                    defaults={'nombre_empresa': fila[1], 'estado': fila[2]}
                )
                ClienteDatos.objects.update_or_create(
                    cliente=cliente,
                    defaults={
                        'direccion': fila[4],
                        'telefono': fila[5],
                        'correo': fila[6],
                        'pagina_web': fila[7],
                        'numero_sucursales': fila[8],
                        'numero_empleados': fila[9],
                        'principales_productos': fila[10],
                    }
                )
        return redirect('importar_home')

@csrf_exempt
@login_required
def importar_clientes_contactos_excel(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo = request.POST.get('modo', 'columnas')
        if archivo:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            for fila in ws.iter_rows(min_row=2, values_only=True):
                id_zoho = fila[0]
                if not id_zoho:
                    continue
                cliente, _ = Cliente.objects.update_or_create(
                    id_zoho=id_zoho,
                    defaults={'nombre_empresa': fila[1], 'estado': fila[2]}
                )
                if modo == 'filas':
                    ClienteContacto.objects.update_or_create(
                        cliente=cliente,
                        correo=fila[6],
                        defaults={
                            'nombre': fila[4].split()[0] if fila[4] else '',
                            'apellido': ' '.join(fila[4].split()[1:]) if fila[4] else '',
                            'telefono': fila[5],
                            'cargo': ''
                        }
                    )
                else:
                    for i in range(4, 16, 3):
                        if fila[i]:
                            ClienteContacto.objects.update_or_create(
                                cliente=cliente,
                                correo=fila[i+2],
                                defaults={
                                    'nombre': fila[i].split()[0] if fila[i] else '',
                                    'apellido': ' '.join(fila[i].split()[1:]) if fila[i] else '',
                                    'telefono': fila[i+1],
                                    'cargo': ''
                                }
                            )
        return redirect('importar_home')


# --- NUEVAS VISTAS: resumen_asignacion_json y asignar_usuarios_pendientes ---
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q

@login_required
def resumen_asignacion_json(request):
    grupo_colectores = Group.objects.filter(name='colector').first()
    if not grupo_colectores:
        return JsonResponse({'usuarios': [], 'sin_asignar': 0})

    usuarios = grupo_colectores.user_set.all()
    resumen = []
    for u in usuarios:
        asignados = Cliente.objects.filter(usuario_asignado=u).count()
        actualizados = Cliente.objects.filter(usuario_asignado=u, estado='Actualizado').count()
        no_ubicados = Cliente.objects.filter(usuario_asignado=u, estado='No Ubicado').count()
        pendientes = Cliente.objects.filter(usuario_asignado=u, estado='Pendiente').count()
        resumen.append({
            'nombre': f"{u.first_name} {u.last_name}",
            'asignados': asignados,
            'actualizados': actualizados,
            'no_ubicados': no_ubicados,
            'pendientes': pendientes,
        })

    sin_asignar = Cliente.objects.filter(usuario_asignado__isnull=True, estado='Pendiente').count()

    return JsonResponse({'usuarios': resumen, 'sin_asignar': sin_asignar})


@login_required
@require_http_methods(["POST"])
def asignar_usuarios_pendientes(request):
    usuario_id = request.POST.get('usuario_id')
    cantidad = int(request.POST.get('cantidad', 0))

    if usuario_id and cantidad > 0:
        usuario = User.objects.filter(id=usuario_id).first()
        if usuario:
            pendientes = Cliente.objects.filter(usuario_asignado__isnull=True, estado='Pendiente')[:cantidad]
            pendientes.update(usuario_asignado=usuario)
    
    return redirect('clientes_admin')

# Vista clientes para admin, renderiza dashboard/clientes.html
from django.shortcuts import render
from django.core.paginator import Paginator

@login_required
def clientes(request):
    user_groups = request.user.groups.values_list('name', flat=True)
    tab = request.GET.get('tab', 'pendientes')
    query = request.GET.get('q', '')  # Búsqueda por nombre 
    
    if tab == 'pendientes':
        clientes_qs = Cliente.objects.filter(estado='Pendiente')
    elif tab == 'actualizados':
        clientes_qs = Cliente.objects.filter(estado='Actualizado')
    elif tab == 'no_ubicados':
        clientes_qs = Cliente.objects.filter(estado='No Ubicado')
    else:
        clientes_qs = Cliente.objects.none()

    if query:
        clientes_qs = clientes_qs.filter(nombre_empresa__icontains=query)

    # Paginación
    paginator = Paginator(clientes_qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'is_admin': request.user.is_superuser or 'admin' in user_groups,
        'is_colector': 'colector' in user_groups,
        'page_obj': page_obj,
        'tab': tab,
        'query': query,  # Para mantener la búsqueda en el input
    }

    # --- ASIGNACIÓN MASIVA: usuarios_colectores y resumen_asignaciones ---
    from django.db.models import Count, Q
    from django.contrib.auth import get_user_model

    User = get_user_model()
    usuarios_colectores = User.objects.filter(groups__name='colector')

    resumen_asignaciones = []
    for u in usuarios_colectores:
        resumen_asignaciones.append({
            'usuario': u,
            'total': Cliente.objects.filter(usuario_asignado=u).count(),
            'pendientes': Cliente.objects.filter(usuario_asignado=u, estado='Pendiente').count(),
            'actualizados': Cliente.objects.filter(usuario_asignado=u, estado='Actualizado').count(),
            'no_ubicados': Cliente.objects.filter(usuario_asignado=u, estado='No Ubicado').count(),
        })

    context['usuarios_colectores'] = usuarios_colectores
    context['resumen_asignaciones'] = resumen_asignaciones

    return render(request, 'dashboard/clientes.html', context)

# --- NUEVA VISTA: Asignar Clientes Masivamente ---
from django.contrib import messages
from django.contrib.auth import get_user_model

@login_required
def asignar_clientes_masivo(request):
    if request.method == 'POST':
        user_id = request.POST.get('usuario_id')
        cantidad = int(request.POST.get('cantidad', 0))

        try:
            usuario = User.objects.get(id=user_id)
            clientes_a_asignar = list(
                Cliente.objects.filter(estado='Pendiente', usuario_asignado__isnull=True)
                .order_by('id')[:cantidad]
            )

            for cliente in clientes_a_asignar:
                cliente.usuario_asignado = usuario
                cliente.save()

            messages.success(request, f'Se asignaron {len(clientes_a_asignar)} clientes a {usuario.get_full_name()}')
        except User.DoesNotExist:
            messages.error(request, 'Usuario no encontrado')
        except Exception as e:
            messages.error(request, f'Error al asignar: {str(e)}')

    return redirect('clientes_admin')


from django.contrib import messages

# --- NUEVA VISTA: Carga Inicial de Clientes ---
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

@login_required
def carga_inicial_clientes(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        df = pd.read_excel(archivo)

        ClienteContacto.objects.all().delete()
        ClienteDatos.objects.all().delete()
        Cliente.objects.all().delete()

        def dividir_nombre_apellido(nombre_completo):
            partes = str(nombre_completo).strip().split()
            if len(partes) >= 3:
                return partes[0], " ".join(partes[1:])
            elif len(partes) == 2:
                return partes[0], partes[1]
            elif len(partes) == 1:
                return partes[0], ""
            else:
                return "", ""

        for _, row in df.iterrows():
            cliente = Cliente.objects.create(
                id_zoho=str(row['id_zoho']).strip(),
                nombre_empresa=str(row['nombre_empresa']).strip(),
                estado=str(row['estado']).strip(),
                observaciones='',
            )

            ClienteDatos.objects.create(
                cliente=cliente,
                direccion=str(row['direccion']).strip() if pd.notna(row['direccion']) else '',
                telefono=str(row['telefono']).strip() if pd.notna(row['telefono']) else '',
                correo=str(row['correo']).strip() if pd.notna(row['correo']) else '',
            )

            nombre, apellido = dividir_nombre_apellido(row['nombre_apellido_contacto'])
            ClienteContacto.objects.create(
                cliente=cliente,
                nombre=nombre,
                apellido=apellido,
                telefono=str(row['telefono']).strip() if pd.notna(row['telefono']) else '',
                correo=str(row['correo']).strip() if pd.notna(row['correo']) else '',
                cargo=str(row['cargo']).strip() if pd.notna(row['cargo']) else '',
            )

        messages.success(request, 'Importación de clientes realizada con éxito.')
        return redirect('importar_clientes')

    return render(request, 'dashboard/importar.html')

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect

@login_required
def reasignar_cliente(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        usuario_id = request.POST.get('usuario_id')

        if not cliente_id or not usuario_id:
            messages.error(request, 'Faltan datos para reasignar el cliente.')
            return redirect('dashboard:clientes')

        try:
            cliente = get_object_or_404(Cliente, id=int(cliente_id))
            usuario = get_object_or_404(User, id=int(usuario_id))
            cliente.usuario_asignado = usuario
            cliente.save()
            messages.success(request, 'Cliente reasignado exitosamente.')
        except ValueError:
            messages.error(request, 'Error al procesar los datos.')

        return redirect('clientes_admin')